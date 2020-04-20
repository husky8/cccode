import xlrd
import xlwt
from xlutils.copy import copy
import openpyxl

class GetInfoFromExcel():
    @staticmethod
    def getInfoFromExcel(filePath, sheetName=None, sheetIndex=None, row=None, col=None):
        """给定EXCEL文件以及sheet名称或索引返回对应 行或列或整张表的**"""
        workBook = xlrd.open_workbook(filePath)
        if sheetName is None and sheetIndex is None:
            sheet = workBook.sheet_by_index(0)
        elif sheetName is not None:
            try:
                sheet = workBook.sheet_by_name(sheetName)
            except xlrd.biffh.XLRDError:
                return "给定的sheetName不存在"

        else:
            try:
                sheet = workBook.sheet_by_index(sheetIndex)
            except IndexError:
                return "给定的sheetIndex超限"
        if row is None and col is None:
            resList = []
            for i in range(sheet.nrows):
                resList.append(sheet.row_values(i))
            return resList
        if row is not None and col is not None:
            return sheet.cell(row, col).value
        if row is not None:
            return sheet.row_values(row)
        if col is not None:
            return sheet.col_values(col)
    @staticmethod
    def getCountFromExcel(filePath, sheetName=None, sheetIndex=None):
        """给定EXCEL文件以及sheet名称或索引返回对应 (行数，列数）"""
        try:
            workBook = xlrd.open_workbook(filePath)
            if sheetName is None and sheetIndex is None:
                sheet = workBook.sheet_by_index(0)
            elif sheetName is not None:
                try:
                    sheet = workBook.sheet_by_name(sheetName)
                except xlrd.biffh.XLRDError:
                    return "给定的sheetName不存在"

            else:
                try:
                    sheet = workBook.sheet_by_index(sheetIndex)
                except IndexError:
                    return "给定的sheetIndex超限"
            cols = sheet.ncols
            rows = sheet.nrows
            return (rows,cols)
        except:
            return ("获取行数失败","获取列数失败")



class WriteDataToExcel():
    @staticmethod
    #nouse
    def writeDataToExcel_v0( fileName,datass, sheetIndex=0,备注="暂只支持行尾追加"):
        """
        给定数据格式应为二维数组  [[a1 , b1, c1],[a2 , b2, c2],[a3 , b3, c3],]
        数据将被解析为 第一行 [a1 , b1, c1] 第二行 [a2 , b2, c2] 第三行 [a3 , b3, c3]
        给定EXCEL文件以及sheet名称或索引返回对应 行或列或整张表的**"""
        rb = xlrd.open_workbook(fileName)
        beforeRow = rb.sheet_by_index(sheetIndex).nrows
        wb = copy(rb)
        s = wb.get_sheet(sheetIndex)
        # row=beforeRow+1
        row = beforeRow #因为索引与个数起始数不同 不需要加一
        for datas in datass:
            col = 0
            for data in datas:
                try:
                    if len(data):data=data[:32000]
                    s.write(row,col,data)
                    col = col + 1
                except:
                    print(datas[1])
                    break
            row = row +1
        wb.save(fileName)

    @staticmethod
    def writeDataToExcel( fileName,datass, sheetIndex=0,备注="暂只支持行尾追加"):
        """给定数据格式应为二维数组  [[a1 , b1, c1],[a2 , b2, c2],[a3 , b3, c3],]
                数据将被解析为 第一行 [a1 , b1, c1] 第二行 [a2 , b2, c2] 第三行 [a3 , b3, c3]
                给定EXCEL文件以及sheet名称或索引返回对应 行或列或整张表的**"""
        wb = openpyxl.load_workbook(fileName)
        ws = wb.active
        for datas in datass:
            try:
                ws.append(datas)
            except Exception as e:
                print("{} 写入失败".format(datas[1]))
        wb.save(fileName)
        print("write finished")
        return "finish"

class CreateNewWorkbook():
    @staticmethod
    def createNewWorkbook(filePath,cols=None,sheetName="sheet1"):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(cols)
        wb.save(filePath)
        print("create finish")
        return "create finish"

if __name__ == '__main__':
    l = [['title', 'subtitle', 'remark', 'websiteid', 'websitename', 'websitemallid', 'websitedbtype', 'websitedbtype', 'websitedbtypename', 'websitecurrencytypeid', 'userid', 'groupid', 'groupname', 'grouptypeid', 'systemgroupid', 'grouptypename', 'url', 'basecategoryid1', 'basecategoryid2', 'basecategoryid3', 'basecategoryname1', 'basecategoryname2', 'basecategoryname3', 'price', 'priceremark', 'priceremarkdescription', 'rmbprice', 'finalrmbprice', 'imageurl', 'lowestprice', 'lowestpricechgrate', 'warningprice', 'recommprice', 'recommrmbprice', 'recommcurrencytypeid', 'recommurl', 'recommdate', 'reviewqty', 'goodreviewrate', 'ignore', 'ignorereason', 'ignoredate', 'updateduserid', 'brandid', 'brandname', 'brandrankid', 'wikiid', 'wikiurl', 'wikilevelid', 'autopushreason', 'shopname', 'originalprice', 'originalpricereducerate', 'isthirdpartysaler', 'datasourceid', 'originalnumber', 'vipprice', 'couponprice', 'mdiscountedprice', 'pricepromotion', 'vippricepromotion', 'couponpricepromotion', 'mdiscountedpricepromotion', 'allpricepromotion', 'islowestallchannel', 'avgprice30days', 'avgprice30dayschgrate', 'lowestpricechgrate', 'lowestprice30dayschgrate', 'articlecountinjingxuan', 'articlecountallchanel', 'articlecountinjingxuanrate', 'articleproducttitle', 'finalprice', 'originalfinalprice', 'originalfinalpricereducerate', 'articlecollectioncount', 'articlecommentcount', 'articleworthy', 'articleunworthy', 'articleclicks', 'articlepv', 'skuid', 'skuname', 'futureprice', 'futurepriceremark', 'commissionrate'], ['ARMANI+EXCHANGE+%E9%98%BF%E7%8E%9B%E5%B0%BC%E5%A5%A2%E4%BE%88%E5%93%81%E5%A5%B3%E5%A3%AB%E8%BF%9E%E8%A1%A3%E8%A3%99+3GYA61-YNKHZ+REDWHITE-6401+6', '', ' ', '1', '%E4%BA%AC%E4%B8%9C%E5%95%86%E5%9F%8E', '183', '1', 'D1', '%E5%9B%BD%E5%86%85%E5%95%86%E5%93%81%E9%87%87%E9%9B%861+%28D1%29', '0', '0', '69', '%E5%8F%91%E7%8E%B0', '2', '2', '%E5%AE%A2%E6%88%B7%E7%AB%AF%E5%8D%95%E5%93%81', 'http%3A%2F%2Fitem.jd.com%2F100004552366.html', '57', '59', '1665', ' ', ' ', ' ', '270.0', '%7B%22TypeID%22%3A0%2C%22SourceType%22%3A%22weiXin%22%2C%22SourceTypeName%22%3A%22%E5%BE%AE%E4%BF%A1%22%2C%22Promotion%22%3Anull%7D', '%E5%BE%AE%E4%BF%A1%E4%B8%93%E4%BA%AB', '270.0', '232.5', 'http%3A%2F%2Fimg11.360buyimg.com%2Fn1%2Fs450x450_jfs%2Ft1%2F30776%2F2%2F13196%2F183867%2F5cb98b64E476e57ca%2Fc748793f985f6e50.jpg', '450.0', '0.0', '0.0', '0.0', '0.0', '0', ' ', ' ', '12', '1.0', 'false', ' ', ' ', '0', '989', ' ', '0', '0', ' ', '0', '%E5%8E%86%E5%8F%B2%E6%96%B0%E4%BD%8E', ' ', '900.0', '0.7', 'false', '2', '100004552366', '0.0', '0.0', '232.5', '%5B%5D', ' ', 'null', '%5B%7B%22Detail%22%3A%7B%22MaxRebate%22%3A2000.0%2C%22NeedAmount%22%3A1000.0%2C%22Rebate%22%3A150.0%2C%22SaveTotalMoney%22%3A150.0%2C%22maxRebate%22%3A2000.0%2C%22needAmount%22%3A1000.0%2C%22rebate%22%3A150.0%2C%22saveTotalMoney%22%3A150.0%7D%2C%22Type%22%3A%22MMJ%22%2C%22TypeID%22%3A3%2C%22detail%22%3A%7B%22%24ref%22%3A%22%24%5B0%5D.Detail%22%7D%2C%22type%22%3A%22MMJ%22%2C%22typeID%22%3A3%2C%22typeName%22%3A%22%E6%AF%8F%E6%BB%A1%E5%87%8F%22%7D%5D', '%7B%22DeviceTypeID%22%3A0%2C%22SkuId%22%3Anull%2C%22SkuName%22%3Anull%2C%22PD1%22%3Anull%2C%22PD2%22%3Anull%2C%22PD3%22%3A%5B%7B%22NeedAmount%22%3A1000.0%2C%22Rebate%22%3A150.0%2C%22MaxRebate%22%3A2000.0%2C%22SaveTotalMoney%22%3A150.0%7D%5D%2C%22PD4%22%3Anull%2C%22PD5%22%3Anull%2C%22PD6%22%3Anull%2C%22PD7%22%3Anull%2C%22PD8%22%3Anull%2C%22PD9%22%3Anull%2C%22PD10%22%3Anull%2C%22PD11%22%3Anull%2C%22PD12%22%3Anull%2C%22PD13%22%3Anull%2C%22PD15%22%3Anull%2C%22PD16%22%3Anull%7D', 'false', '1149.63', '0.7977610187625584', '899.0', '0.7413793103448276', '0', '0', '0.0', ' ', '232.5', '825.0', '0.72', '0', '0', '0', '0', '0', '0', ' ', ' ', '0.0', ' ', '0.0'], ['%E5%B0%8F%E7%99%BD%E7%86%8A+%E7%94%B5%E5%8A%A8%E5%90%B8%E5%A5%B6%E5%99%A8+%E9%94%82%E7%94%B5%E6%B1%A0%E5%8F%AF%E5%85%85%E7%94%B5%E5%BC%8F%E5%90%B8%E4%B9%B3%E5%99%A8+%E9%9D%99%E9%9F%B3%E4%BE%BF%E6%90%BA%E6%8B%94%E5%A5%B6%E5%99%A8+%E5%BE%85%E4%BA%A7%E5%8C%85', '', ' ', '1', '%E4%BA%AC%E4%B8%9C%E5%95%86%E5%9F%8E', '183', '1', 'D1', '%E5%9B%BD%E5%86%85%E5%95%86%E5%93%81%E9%87%87%E9%9B%861+%28D1%29', '0', '0', '69', '%E5%8F%91%E7%8E%B0', '2', '2', '%E5%AE%A2%E6%88%B7%E7%AB%AF%E5%8D%95%E5%93%81', 'http%3A%2F%2Fitem.jd.com%2F7590189.html', '75', '83', '913', ' ', ' ', ' ', '199.0', '%7B%22TypeID%22%3A0%2C%22SourceType%22%3A%22weiXin%22%2C%22SourceTypeName%22%3A%22%E5%BE%AE%E4%BF%A1%22%2C%22Promotion%22%3Anull%7D', '%E5%BE%AE%E4%BF%A1%E4%B8%93%E4%BA%AB', '199.0', '189.0', 'http%3A%2F%2Fimg11.360buyimg.com%2Fn1%2Fs450x450_jfs%2Ft1%2F86722%2F33%2F2856%2F124510%2F5dd783deEa2cd9690%2F0bda607c4d731a3b.jpg', '199.0', '0.0', '0.0', '0.0', '0.0', '0', ' ', ' ', '40228', '0.97', 'false', ' ', ' ', '0', '3701', ' ', '0', '0', ' ', '0', '%E5%8E%86%E5%8F%B2%E6%96%B0%E4%BD%8E', ' ', '219.0', '0.09', 'false', '2', '7590189', '0.0', '189.0', '189.0', '%5B%5D', ' ', '%5B%7B%22Detail%22%3A%7B%22CouponClass%22%3A%22DQ%22%2C%22CouponKind%22%3A%22XPL%22%2C%22CouponType%22%3A1%2C%22Discount%22%3A0.0%2C%22EndTime%22%3A1588262399000%2C%22MaxRebate%22%3A0.0%2C%22NeedAmount%22%3A199.0%2C%22Rebate%22%3A10.0%2C%22SaveTotalMoney%22%3A10.0%2C%22StartTime%22%3A1586188800000%2C%22Url%22%3A%22https%3A%2F%2Fcoupon.jd.com%2Filink%2FcouponSendFront%2Fsend_index.action%3Fkey%3D200aadd41dfc4dae9164a74d30c6ef0a%26roleId%3D29591872%26to%3Dhttps%3A%2F%2Fpro.jd.com%2Fmall%2Factive%2F3iu3G9t2BW4tkY3dxkSsBh8RdQk7%2Findex.html%2Chttps%3A%2F%2Fpro.m.jd.com%2Fmall%2Factive%2F3iu3G9t2BW4tkY3dxkSsBh8RdQk7%2Findex.html%22%2C%22couponClass%22%3A%22DQ%22%2C%22couponKind%22%3A%22XPL%22%2C%22discount%22%3A0.0%2C%22endTime%22%3A1588262399000%2C%22maxRebate%22%3A0.0%2C%22needAmount%22%3A199.0%2C%22numCouponClass%22%3A1%2C%22numCouponKind%22%3A1%2C%22rebate%22%3A10.0%2C%22saveTotalMoney%22%3A10.0%2C%22startTime%22%3A1586188800000%2C%22url%22%3A%22https%3A%2F%2Fcoupon.jd.com%2Filink%2FcouponSendFront%2Fsend_index.action%3Fkey%3D200aadd41dfc4dae9164a74d30c6ef0a%26roleId%3D29591872%26to%3Dhttps%3A%2F%2Fpro.jd.com%2Fmall%2Factive%2F3iu3G9t2BW4tkY3dxkSsBh8RdQk7%2Findex.html%2Chttps%3A%2F%2Fpro.m.jd.com%2Fmall%2Factive%2F3iu3G9t2BW4tkY3dxkSsBh8RdQk7%2Findex.html%22%7D%2C%22Type%22%3A%22LQ%22%2C%22TypeID%22%3A7%2C%22detail%22%3A%7B%22%24ref%22%3A%22%24%5B0%5D.Detail%22%7D%2C%22type%22%3A%22LQ%22%2C%22typeID%22%3A7%2C%22typeName%22%3A%22%E9%A2%86%E5%88%B8%22%7D%5D', '%5B%7B%22Detail%22%3A%7B%22CouponClass%22%3A%22DQ%22%2C%22CouponKind%22%3A%22XPL%22%2C%22CouponType%22%3A1%2C%22Discount%22%3A0.0%2C%22EndTime%22%3A1588262399000%2C%22MaxRebate%22%3A0.0%2C%22NeedAmount%22%3A199.0%2C%22Rebate%22%3A10.0%2C%22SaveTotalMoney%22%3A10.0%2C%22StartTime%22%3A1586188800000%2C%22Url%22%3A%22https%3A%2F%2Fcoupon.jd.com%2Filink%2FcouponSendFront%2Fsend_index.action%3Fkey%3D200aadd41dfc4dae9164a74d30c6ef0a%26roleId%3D29591872%26to%3Dhttps%3A%2F%2Fpro.jd.com%2Fmall%2Factive%2F3iu3G9t2BW4tkY3dxkSsBh8RdQk7%2Findex.html%2Chttps%3A%2F%2Fpro.m.jd.com%2Fmall%2Factive%2F3iu3G9t2BW4tkY3dxkSsBh8RdQk7%2Findex.html%22%2C%22couponClass%22%3A%22DQ%22%2C%22couponKind%22%3A%22XPL%22%2C%22discount%22%3A0.0%2C%22endTime%22%3A1588262399000%2C%22maxRebate%22%3A0.0%2C%22needAmount%22%3A199.0%2C%22numCouponClass%22%3A1%2C%22numCouponKind%22%3A1%2C%22rebate%22%3A10.0%2C%22saveTotalMoney%22%3A10.0%2C%22startTime%22%3A1586188800000%2C%22url%22%3A%22https%3A%2F%2Fcoupon.jd.com%2Filink%2FcouponSendFront%2Fsend_index.action%3Fkey%3D200aadd41dfc4dae9164a74d30c6ef0a%26roleId%3D29591872%26to%3Dhttps%3A%2F%2Fpro.jd.com%2Fmall%2Factive%2F3iu3G9t2BW4tkY3dxkSsBh8RdQk7%2Findex.html%2Chttps%3A%2F%2Fpro.m.jd.com%2Fmall%2Factive%2F3iu3G9t2BW4tkY3dxkSsBh8RdQk7%2Findex.html%22%7D%2C%22Type%22%3A%22LQ%22%2C%22TypeID%22%3A7%2C%22detail%22%3A%7B%22%24ref%22%3A%22%24%5B0%5D.Detail%22%7D%2C%22type%22%3A%22LQ%22%2C%22typeID%22%3A7%2C%22typeName%22%3A%22%E9%A2%86%E5%88%B8%22%7D%5D', '%7B%22DeviceTypeID%22%3A0%2C%22SkuId%22%3Anull%2C%22SkuName%22%3Anull%2C%22PD1%22%3Anull%2C%22PD2%22%3Anull%2C%22PD3%22%3Anull%2C%22PD4%22%3Anull%2C%22PD5%22%3Anull%2C%22PD6%22%3Anull%2C%22PD7%22%3A%5B%7B%22NeedAmount%22%3A199.0%2C%22Rebate%22%3A10.0%2C%22Url%22%3A%22https%3A%2F%2Fcoupon.jd.com%2Filink%2FcouponSendFront%2Fsend_index.action%3Fkey%3D200aadd41dfc4dae9164a74d30c6ef0a%26roleId%3D29591872%26to%3Dhttps%3A%2F%2Fpro.jd.com%2Fmall%2Factive%2F3iu3G9t2BW4tkY3dxkSsBh8RdQk7%2Findex.html%2Chttps%3A%2F%2Fpro.m.jd.com%2Fmall%2Factive%2F3iu3G9t2BW4tkY3dxkSsBh8RdQk7%2Findex.html%22%2C%22StartTime%22%3A%222020-04-07T00%3A00%3A00%22%2C%22EndTime%22%3A%222020-04-30T23%3A59%3A59%22%2C%22Discount%22%3A0.0%2C%22CouponType%22%3A1%2C%22MaxRebate%22%3A0.0%2C%22SaveTotalMoney%22%3A10.0%2C%22IsInsideCoupon%22%3Anull%2C%22SourceType%22%3Anull%2C%22CouponPlus%22%3Anull%2C%22CouponLapDesc%22%3Anull%2C%22CouponKind%22%3A1%2C%22CouponClass%22%3A1%7D%5D%2C%22PD8%22%3Anull%2C%22PD9%22%3Anull%2C%22PD10%22%3Anull%2C%22PD11%22%3Anull%2C%22PD12%22%3Anull%2C%22PD13%22%3Anull%2C%22PD15%22%3Anull%2C%22PD16%22%3Anull%7D', 'false', '206.14', '0.0831473755700009', '199.0', '0.05025125628140704', '0', '0', '0.0', ' ', '189.0', '219.0', '0.14', '0', '0', '0', '0', '0', '0', ' ', ' ', '0.0', ' ', '0.0']]

    # CreateNewWorkbook().createNewWorkbook("11.xlsx",l[0])
    # c = GetInfoFromExcel().getInfoFromExcel("测试用excel.xlsx", sheetName=4)
    # print(c)
    # c = CreateNewWorkbook().createNewWorkbook("ceshi.xls",[1,2,3,4,5,6,7,8,9])
    # print(c)
    import urllib.parse
    ll = []
    for i in [1,2]:
        lll = [urllib.parse.unquote(this) for this in l[i]]
        ll.append(lll)

    WriteDataToExcel().writeDataToExcel("11.xlsx",ll)