import xlrd
from xlutils.copy import copy
import openpyxl
from pathlib import Path

class GetInfoFromExcel():
    @staticmethod
    def getInfoFromExcel(filePath, sheetName=None, sheetIndex=None, row=None, col=None):
        """给定EXCEL文件以及sheet名称或索引返回对应 行或列或整张表的**"""
        workBook = xlrd.open_workbook(filePath)
        if sheetName is None and sheetIndex is None:
            sheet = workBook.sheet_by_index(0)
        elif sheetName is not None:
            try:
                sheet = workBook.sheet_by_name(sheetName)
            except xlrd.biffh.XLRDError:
                return "给定的sheetName不存在"

        else:
            try:
                sheet = workBook.sheet_by_index(sheetIndex)
            except IndexError:
                return "给定的sheetIndex超限"
        if row is None and col is None:
            resList = []
            for i in range(sheet.nrows):
                resList.append(sheet.row_values(i))
            return resList
        if row is not None and col is not None:
            return sheet.cell(row, col).value
        if row is not None:
            return sheet.row_values(row)
        if col is not None:
            return sheet.col_values(col)
    @staticmethod
    def getCountFromExcel(filePath, sheetName=None, sheetIndex=None):
        """给定EXCEL文件以及sheet名称或索引返回对应 (行数，列数）"""
        try:
            workBook = xlrd.open_workbook(filePath)
            if sheetName is None and sheetIndex is None:
                sheet = workBook.sheet_by_index(0)
            elif sheetName is not None:
                try:
                    sheet = workBook.sheet_by_name(sheetName)
                except xlrd.biffh.XLRDError:
                    return "给定的sheetName不存在"

            else:
                try:
                    sheet = workBook.sheet_by_index(sheetIndex)
                except IndexError:
                    return "给定的sheetIndex超限"
            cols = sheet.ncols
            rows = sheet.nrows
            return (rows,cols)
        except:
            return ("获取行数失败","获取列数失败")

    @staticmethod
    def getInfoFromExcelbyOpen(filePath, sheetName=None, sheetIndex=None, row=None, col=None):
        """给定EXCEL文件以及sheet名称或索引返回对应 行或列或整张表的**"""
        workBook = openpyxl.load_workbook(filePath)
        if sheetName is None and sheetIndex is None:
            sheet = workBook.sheet_by_index(0)
        elif sheetName is not None:
            try:
                sheet = workBook.sheet_by_name(sheetName)
            except xlrd.biffh.XLRDError:
                return "给定的sheetName不存在"

        else:
            try:
                sheet = workBook.sheet_by_index(sheetIndex)
            except IndexError:
                return "给定的sheetIndex超限"
        if row is None and col is None:
            resList = []
            for i in range(sheet.nrows):
                resList.append(sheet.row_values(i))
            return resList
        if row is not None and col is not None:
            return sheet.cell(row, col).value
        if row is not None:
            return sheet.row_values(row)
        if col is not None:
            return sheet.col_values(col)


class WriteDataToExcel():
    @staticmethod
    def writeDataToExcel( fileName,datass, sheetIndex=0,备注="暂只支持行尾追加"):
        """给定数据格式应为二维数组  [[a1 , b1, c1],[a2 , b2, c2],[a3 , b3, c3],]
                数据将被解析为 第一行 [a1 , b1, c1] 第二行 [a2 , b2, c2] 第三行 [a3 , b3, c3]
                给定EXCEL文件以及sheet名称或索引返回对应 行或列或整张表的**"""
        if not Path(fileName).exists():
            CreateNewWorkbook().createNewWorkbook(fileName)
        wb = openpyxl.load_workbook(fileName)
        ws = wb.active
        for datas in datass:
            try:
                ws.append(datas)
            except Exception as e:
                print("{} 写入失败".format(datas[1]))
        wb.save(fileName)
        print("write finished")
        return "finish"

class CreateNewWorkbook():
    @staticmethod
    def createNewWorkbook(filePath,cols=None,sheetName="sheet1"):
        wb = openpyxl.Workbook()
        ws = wb.active
        if cols is not None:
            ws.append(cols)
        wb.save(filePath)
        print("create finish")
        return "create finish"

class ModifyExcel():
    @staticmethod
    def modifyExcel(filePath,Coordinate,newValue,sheetName=None):
        wb = openpyxl.load_workbook(filePath)
        if sheetName :
            ws = wb[sheetName]
        else:
            ws = wb.active
        ws[Coordinate]=newValue
        wb.save(filePath)
if __name__ == '__main__':
    WriteDataToExcel().writeDataToExcel("12.xlsx",[[1,2,3,4,5]])

    pass